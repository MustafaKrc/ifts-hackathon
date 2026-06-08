"""
Jira Task Creator/Updater using Atlassian REST API.
Reads tasks from AI_Create_Task.xlsx and creates OR updates Jira issues.

Usage:
    python jira_create_tasks.py <sprint_no>

Column mapping (AI_Create_Task.xlsx):
    A  - Sprint No  → Sprint (only rows matching <sprint_no> are processed)
    B  - Jira No    → Issue key (empty = create, filled = update)
    F  - Ekip       → Labels
    G  - Jira Adı   → Summary
    H  - Madde      → Description
    I  - SP         → Story Points
    J  - ICT Kişi   → Assignee
    M  - Proje      → Epic Link

Rules:
    - Only rows where column A == <sprint_no> are processed.
    - 'Jira No' EMPTY  → Create new task, write key back to Excel.
    - 'Jira No' FILLED → Fetch existing task and update fields if different.
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace', line_buffering=True)
import argparse
import os
import openpyxl
import requests
import json
from requests.auth import HTTPBasicAuth
from dotenv import load_dotenv

# .env dosyasini yukle (scriptle ayni klasorde olmali)
load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), ".env"))

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────
JIRA_URL   = os.environ["JIRA_URL"]
EMAIL      = os.environ["JIRA_EMAIL"]
API_TOKEN  = os.environ["JIRA_API_TOKEN"]

# AUTH_MODE:
#   "bearer" → Jira Server/Data Center Personal Access Token (PAT)
#   "basic"  → Atlassian Cloud (email + API token)
AUTH_MODE  = os.environ.get("JIRA_AUTH_MODE", "bearer")
print("DEBUG JIRA_URL:", JIRA_URL)
print("DEBUG JIRA_EMAIL:", EMAIL)
print("DEBUG AUTH_MODE:", AUTH_MODE)
print("DEBUG TOKEN_LENGTH:", len(API_TOKEN))

EXCEL_FILE  = r"C:\Users\TCKUYUCEL\Desktop\PLanlama\AI_Create_Task.xlsx"
SHEET_NAME  = None   # None = use first sheet
ISSUE_TYPE  = "Task"

# ── Column positions (1-based) ────────────────────────────────────────────────
COL_SPRINT   = 1   # A - Sprint No
COL_JIRA_NO  = 2   # B - Jira No
COL_SQUAD    = 5   # E - Squad      → Project Key (US / POS)
COL_EKIP     = 6   # F - Ekip       → Labels
COL_SUMMARY  = 7   # G - Jira Adı   → Summary
COL_DESC     = 8   # H - Madde      → Description
COL_SP       = 9   # I - SP         → Story Points
COL_ASSIGNEE = 10  # J - ICT Kişi   → Assignee
COL_PROJE    = 13  # M - Proje      → Epic Link
# ─────────────────────────────────────────────

# ── Assignee: project key → (Excel name (lowercase) → Jira username) ────────
USER_MAP = {
    "US": {
        "anıl"    : "TCAPALABIYIK",
        "anil"    : "TCAPALABIYIK",
        "aziz"    : "TCAOKAY",
        "burcu"   : "TCBUKARAAGAC",
        "derya"   : "TCDGUMUSSOY",
        "erhan"   : "TCEGEZINTI",
        "hamza"   : "TCHCOLAK",
        "hazal"   : "TCHUKECOGLU",
        "ismail"  : "TCMUYETKIN",
        "i\u0307smail" : "TCMUYETKIN",   # Turkish İ → i with combining dot
        "koray ç.": "TCKCELIKBAS",
        "koray c.": "TCKCELIKBAS",
        "koray"   : "TCKKURBAN",
        "sinem"   : "TCSIUZUN",
        "soner"   : "TCSTAKATAS",
    },
    "POS": {
        "erol"    : "TCETUTUMLU",
        "merve"   : "TCMEROCAK",
        "mustafa" : "TCMKIRCI",
        "oğuz"    : "TCOGTOPAL",
        "oguz"    : "TCOGTOPAL",
        "osman"   : "TCOZAIM",
        "özgül"   : "TCOERDEBIL",
        "ozgul"   : "TCOERDEBIL",
        "sinem"   : "TCSPATAR",
        "ünal"    : "TCUZAFER",
        "unal"    : "TCUZAFER",
    },
}

# ── Label: Excel 'Ekip' value (lowercase) → Jira label ───────────────────────
LABEL_MAP = {
    "cc"    : "CREDITCONTROL",
    "fraud" : "FRAUD",
    "töhaş" : "TOHAS",
    "tohas" : "TOHAS",
}
# ─────────────────────────────────────────────

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

if AUTH_MODE == "bearer":
    auth    = None
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Authorization": f"Bearer {API_TOKEN}"
    }
else:
    auth    = HTTPBasicAuth(EMAIL, API_TOKEN)
    headers = {"Accept": "application/json", "Content-Type": "application/json"}

SSL_VERIFY = False   # Corporate proxy uses self-signed cert

# Cache: epic name (lowercase) → epic key (e.g. "US-2548")
_epic_key_cache = {}

def resolve_epic_key(epic_name: str, project_key: str) -> str | None:
    """Search Jira for an Epic by name and return its issue key."""
    if not epic_name or not str(epic_name).strip():
        return None

    name = str(epic_name).strip()
    cache_key = name.lower()
    if cache_key in _epic_key_cache:
        return _epic_key_cache[cache_key]

    try:
        jql = f'project = {project_key} AND issuetype = Epic AND "Epic Name" ~ "{name}"'
        url = f"{JIRA_URL}/rest/api/2/search?jql={requests.utils.quote(jql)}&fields=summary,customfield_10030&maxResults=10"
        resp = requests.get(url, auth=auth, headers=headers, verify=SSL_VERIFY, timeout=30)
        resp.raise_for_status()
        issues = resp.json().get("issues", [])

        for issue in issues:
            epic_field = (issue.get("fields") or {}).get("customfield_10030") or ""
            summary    = (issue.get("fields") or {}).get("summary") or ""
            # Match against Epic Name field or summary
            if (epic_field.strip().lower() == name.lower() or
                    summary.strip().lower() == name.lower()):
                key = issue["key"]
                _epic_key_cache[cache_key] = key
                print(f"  [Epic] '{name}' -> {key}")
                return key

        # Fallback: return first result if any
        if issues:
            key = issues[0]["key"]
            _epic_key_cache[cache_key] = key
            print(f"  [Epic] '{name}' -> {key} (closest match)")
            return key

        print(f"  [WARN] Epic '{name}' not found in Jira. Epic Link will not be set.")
        _epic_key_cache[cache_key] = None
        return None
    except Exception as e:
        print(f"  [WARN] Could not search for epic '{name}': {e}")
        return None


def resolve_label(ekip_value: str) -> str | None:
    """Map Excel 'Ekip' value to a Jira label string. Returns None if no match."""
    if not ekip_value or not str(ekip_value).strip():
        return None
    key = str(ekip_value).strip().lower()
    return LABEL_MAP.get(key)   # None if not in map → label stays null


def resolve_assignee(name: str, project_key: str) -> str | None:
    """Map Excel 'ICT Kişi' to a Jira username via USER_MAP."""
    if not name or not str(name).strip():
        return None
    key = str(name).strip().lower()
    mapped = USER_MAP.get(project_key, {}).get(key)
    if not mapped:
        print(f"  [WARN] No USER_MAP entry for '{name}' in project '{project_key}'. Task will be unassigned.")
    return mapped


def get_sprint_id(sprint_name: str, project_key: str) -> int | None:
    """Find the sprint ID by name using the Agile API."""
    try:
        url = f"{JIRA_URL}/rest/agile/1.0/board?projectKeyOrId={project_key}&maxResults=50"
        resp = requests.get(url, auth=auth, headers=headers, verify=SSL_VERIFY, timeout=30)
        resp.raise_for_status()
        boards = resp.json().get("values", [])

        print(f"  [DEBUG] Found {len(boards)} board(s) for project '{project_key}':")
        for b in boards:
            print(f"          Board ID={b['id']}  Name={b.get('name', '')}")

        if not boards:
            print(f"[WARN] No board found for project '{project_key}'. Sprint will not be set.")
            return None

        for board in boards:
            board_id = board["id"]
            start = 0
            while True:
                sprint_url = (
                    f"{JIRA_URL}/rest/agile/1.0/board/{board_id}/sprint"
                    f"?startAt={start}&maxResults=50"
                )
                sresp = requests.get(sprint_url, auth=auth, headers=headers, verify=SSL_VERIFY, timeout=30)
                if sresp.status_code != 200:
                    print(f"  [WARN] Could not fetch sprints for board {board_id}: {sresp.status_code}")
                    break
                data = sresp.json()
                for sprint in data.get("values", []):
                    print(f"          Sprint: ID={sprint['id']}  Name={sprint['name']}")
                    if sprint["name"].strip().lower() == sprint_name.strip().lower():
                        return sprint["id"]
                if data.get("isLast", True):
                    break
                start += 50

        print(f"[WARN] Sprint '{sprint_name}' not found in any board.")
        return None
    except Exception as e:
        print(f"[WARN] Could not fetch sprint ID: {e}. Tasks will be created without sprint.")
        return None


def get_issue(issue_key: str) -> dict:
    """Fetch an existing Jira issue and return its fields."""
    try:
        url  = f"{JIRA_URL}/rest/api/2/issue/{issue_key}"
        resp = requests.get(url, auth=auth, headers=headers, verify=SSL_VERIFY, timeout=30)
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        print(f"  [WARN] Could not fetch issue {issue_key}: {e}")
        return {}


def create_issue(
    summary: str,
    description: str,
    story_points: float | None,
    assignee_id: str | None,
    sprint_id: int | None,
    label: str | None,
    epic_link: str | None,
    project_key: str = "US",
) -> dict:
    """Create a Jira issue and return the response JSON."""
    fields = {
        "project"    : {"key": project_key},
        "summary"    : summary or "(no summary)",
        "issuetype"  : {"name": ISSUE_TYPE},
        "description": str(description or ""),
    }

    if assignee_id:
        fields["assignee"] = (
            {"name": assignee_id} if AUTH_MODE == "bearer"
            else {"accountId": assignee_id}
        )

    if story_points is not None:
        fields["customfield_10028"] = float(story_points)

    if label:
        fields["labels"] = [label]

    if epic_link:
        fields["customfield_10021"] = epic_link   # Epic Link (Jira Server)

    payload = json.dumps({"fields": fields})
    api_version = "2" if AUTH_MODE == "bearer" else "3"
    url  = f"{JIRA_URL}/rest/api/{api_version}/issue"
    resp = requests.post(url, data=payload, auth=auth, headers=headers, verify=SSL_VERIFY, timeout=30)

    if resp.status_code not in (200, 201):
        error_text = resp.text
        # Retry 1: drop Story Points if field is not on screen
        if "customfield_10028" in error_text:
            print("  [WARN] Story points field not on screen, retrying without SP...")
            fields.pop("customfield_10028", None)
            payload = json.dumps({"fields": fields})
            resp = requests.post(url, data=payload, auth=auth, headers=headers, verify=SSL_VERIFY, timeout=30)
            error_text = resp.text
        # Retry 2: drop Epic Link if it's causing an error
        if resp.status_code not in (200, 201) and "customfield_10021" in error_text:
            print("  [WARN] Epic Link field not on screen, retrying without Epic Link...")
            fields.pop("customfield_10021", None)
            payload = json.dumps({"fields": fields})
            resp = requests.post(url, data=payload, auth=auth, headers=headers, verify=SSL_VERIFY, timeout=30)
            error_text = resp.text
        # Retry 3: drop assignee if it's causing an error
        if resp.status_code not in (200, 201) and "assignee" in error_text:
            print("  [WARN] Assignee cannot be set, retrying without assignee...")
            fields.pop("assignee", None)
            payload = json.dumps({"fields": fields})
            resp = requests.post(url, data=payload, auth=auth, headers=headers, verify=SSL_VERIFY, timeout=30)

    if resp.status_code not in (200, 201):
        print(f"[ERROR] Failed to create issue: {resp.status_code} – {resp.text}")
        return {}

    result = resp.json()

    # Assign issue to sprint via Agile API
    if sprint_id and result.get("key"):
        assign_sprint(result["key"], sprint_id)

    return result


def assign_sprint(issue_key: str, sprint_id: int):
    """Move an issue into a sprint using the Jira Agile API."""
    url = f"{JIRA_URL}/rest/agile/1.0/sprint/{sprint_id}/issue"
    payload = json.dumps({"issues": [issue_key]})
    resp = requests.post(url, data=payload, auth=auth, headers=headers, verify=SSL_VERIFY, timeout=30)
    if resp.status_code in (200, 201, 204):
        print(f"  [OK] Sprint : '{issue_key}' added to sprint {sprint_id}")
    else:
        print(f"  [WARN] Could not set sprint for {issue_key}: {resp.status_code} - {resp.text}")


def update_issue(issue_key: str, fields: dict) -> bool:
    """Update an existing Jira issue with the given fields."""
    payload = json.dumps({"fields": fields})
    url  = f"{JIRA_URL}/rest/api/2/issue/{issue_key}"
    resp = requests.put(url, data=payload, auth=auth, headers=headers, verify=SSL_VERIFY, timeout=30)
    if resp.status_code == 204:
        return True
    print(f"  [WARN] Update failed for {issue_key}: {resp.status_code} – {resp.text}")
    return False


def main():
    parser = argparse.ArgumentParser(
        description="Create/update Jira tasks from AI_Create_Task.xlsx."
    )
    parser.add_argument(
        "--excel-path",
        default=None,
        help="Excel dosyasinin yolu (verilmezse sabit yol kullanilir)",
    )
    args = parser.parse_args()

    excel_file = args.excel_path or EXCEL_FILE

    print(f"Excel     : {excel_file}")
    print()

    # ── Load workbook ─────────────────────────────────────────────────────────
    # Check if Excel file is locked (open in another process)
    try:
        with open(excel_file, "r+b"):
            pass
    except PermissionError:
        print(f"[ERROR] '{excel_file}' baska bir program tarafindan acik.")
        print("        Lutfen Excel'i kapatip tekrar calistirin.")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_file)
    ws = wb[SHEET_NAME] if SHEET_NAME else wb.active
    print(f"Sheet     : {ws.title}\n")

    # Sprint ID cache: sprint_name -> sprint_id (avoid repeated API calls)
    sprint_id_cache = {}

    created     = 0
    updated     = 0
    skipped     = 0
    failed_rows = []   # (row_idx, reason)

    for row_idx in range(2, ws.max_row + 1):
        def cell(col_idx):
            return ws.cell(row=row_idx, column=col_idx).value

        # ── Read sprint number from column A, skip empty rows ─────────────────
        row_sprint_val = cell(COL_SPRINT)
        if row_sprint_val is None or str(row_sprint_val).strip() == "":
            continue

        sprint_no   = str(row_sprint_val).strip()
        sprint_name = f"Sprint {sprint_no}"

        # ── Read fields ───────────────────────────────────────────────────────
        project_key = str(cell(COL_SQUAD)    or "").strip().upper()

        # Fetch sprint ID (use cache to avoid duplicate API calls)
        if sprint_name not in sprint_id_cache:
            print(f"Fetching Sprint ID for '{sprint_name}' ...")
            sprint_id_cache[sprint_name] = get_sprint_id(sprint_name, project_key)
            print(f"Sprint ID : {sprint_id_cache[sprint_name]}\n")
        sprint_id = sprint_id_cache[sprint_name]
        jira_no     = str(cell(COL_JIRA_NO)  or "").strip()
        ekip_raw    = cell(COL_EKIP)
        summary     = str(cell(COL_SUMMARY)  or "").strip()
        description = str(cell(COL_DESC)     or "").strip()
        sp_raw      = cell(COL_SP)
        ict_kisi    = str(cell(COL_ASSIGNEE) or "").strip()
        proje_raw   = str(cell(COL_PROJE)    or "").strip()

        try:
            story_points = float(sp_raw) if sp_raw is not None else None
        except (ValueError, TypeError):
            story_points = None

        label       = resolve_label(ekip_raw)
        assignee_id = resolve_assignee(ict_kisi, project_key) if ict_kisi else None
        epic_link   = resolve_epic_key(proje_raw, project_key) if proje_raw else None

        if not summary:
            print(f"[SKIP] Row {row_idx}: empty 'Jira Adı', skipping.")
            skipped += 1
            continue

        # ══════════════════════════════════════════════════════════════════════
        # CASE 1 – Jira No empty → CREATE
        # ══════════════════════════════════════════════════════════════════════
        if not jira_no:
            print(
                f"[CREATE] Row {row_idx} | {ict_kisi or '(unassigned)'} | "
                f"{summary[:50]} | SP={story_points} | Label={label} | Epic={epic_link}"
            )
            result = create_issue(
                summary, description, story_points,
                assignee_id, sprint_id, label, epic_link,
                project_key
            )
            if result.get("key"):
                new_key = result["key"]
                print(f"  [OK] Created : {new_key}  ->  {JIRA_URL}/browse/{new_key}")
                ws.cell(row=row_idx, column=COL_JIRA_NO).value = new_key
                try:
                    wb.save(excel_file)
                    print(f"  [OK] Saved   : '{new_key}' written to Excel row {row_idx}")
                    created += 1
                except PermissionError:
                    print(f"  [WARN] Excel dosyasi acik, '{new_key}' Row {row_idx}'e yazilamadi.")
                    print(f"         Excel'i kapatip tekrar calistirdiginizda bu satir tekrar denenecek.")
                    failed_rows.append((row_idx, f"excel save failed - Jira task olusturuldu: {new_key} ama yazılamadı"))
            else:
                print(f"  [FAIL] Failed to create row {row_idx}")
                failed_rows.append((row_idx, "create failed"))

        # ══════════════════════════════════════════════════════════════════════
        # CASE 2 – Jira No filled → CHECK & UPDATE if different
        # ══════════════════════════════════════════════════════════════════════
        else:
            existing = get_issue(jira_no)
            if not existing:
                skipped += 1
                failed_rows.append((row_idx, f"get_issue failed ({jira_no})"))
                continue

            ex_fields = existing.get("fields", {})

            ex_summary    = str(ex_fields.get("summary")          or "").strip()
            ex_desc       = str(ex_fields.get("description")      or "").strip()
            ex_sp         = ex_fields.get("customfield_10028")
            ex_labels     = ex_fields.get("labels", [])            # list of strings
            ex_epic_link  = ex_fields.get("customfield_10021")     # Epic Link key
            ex_assignee   = (ex_fields.get("assignee") or {}).get("name")
            changes    = {}
            change_log = []

            if ex_summary != summary:
                changes["summary"] = summary
                change_log.append("summary")

            if ex_desc != description:
                changes["description"] = description
                change_log.append("description")

            if story_points is not None and ex_sp != story_points:
                changes["customfield_10028"] = story_points
                change_log.append(f"SP ({ex_sp} → {story_points})")

            # Labels: compare current Jira label with desired label
            desired_labels = [label] if label else []
            if sorted(ex_labels) != sorted(desired_labels):
                changes["labels"] = desired_labels
                change_log.append(f"labels ({ex_labels} → {desired_labels})")

            # Epic Link
            ex_epic_str = str(ex_epic_link or "").strip()
            new_epic_str = str(epic_link or "").strip()
            if ex_epic_str != new_epic_str:
                changes["customfield_10021"] = epic_link if epic_link else None
                change_log.append(f"epic ({ex_epic_str or 'none'} → {new_epic_str or 'none'})")

            # Assignee
            if assignee_id and ex_assignee != assignee_id:
                changes["assignee"] = {"name": assignee_id}
                change_log.append(f"assignee ({ex_assignee} → {assignee_id})")

            if changes:
                print(f"[UPDATE] {jira_no} | Row {row_idx} | Changing: {', '.join(change_log)}")
                if update_issue(jira_no, changes):
                    print(f"  [OK] Updated fields")
                    updated += 1
                else:
                    skipped += 1
                    failed_rows.append((row_idx, f"update failed ({jira_no})"))

            if not changes:
                print(f"[OK]     {jira_no} | Row {row_idx} | No changes needed")

    print(f"\n{'='*60}")
    print(f"Done.  Created: {created}  |  Updated: {updated}  |  Skipped: {skipped}")

    if failed_rows:
        print(f"\nBasarisiz satirlar ({len(failed_rows)}) - retry icin tekrar calistir:")
        for row_idx, reason in failed_rows:
            print(f"  Row {row_idx} - {reason}")
    else:
        print("\nTum satirlar basariyla islendi.")


if __name__ == "__main__":
    main()
